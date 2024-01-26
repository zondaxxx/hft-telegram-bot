from openpyxl import load_workbook

async def fio(text: str):
    if len(text) > 6 and len(text.split()) == 2:
        a, bc  = text.split()
        if len(a) > 2 and len(bc.split('.')) > 2:
            return True
    return False

async def klass(kls: str):
    if kls.isdigit():
        if 1 <= int(kls) <= 11:
            return True
    return False

async def vl_exl(name):
    file = load_workbook("data/tests/" + name + ".xlsx").active
    if file['A1'].value and file['B1'].value is not None:
        return [True, file.max_row]
    return [False]