from openpyxl import load_workbook

async def question_answer(name) -> list:
    file = load_workbook("data/tests/" + name + ".xlsx").active
    return tuple(file.iter_cols(values_only=True))